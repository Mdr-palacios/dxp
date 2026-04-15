# powerbuilder/chat/tests/test_agents.py
"""
Live integration tests for messaging.py, finance_agent.py, and export.py.
All agents are exercised against Virginia's 7th Congressional District (VA-07,
GEOID 5107) using a manually constructed AgentState so no prior pipeline run
is required.

Test sections:
  0. Pre-flight: API key checks
  1. messaging_node — mock precinct demographics + mock research for young voters
  1a. All five section markers parsed from LLM response
  1b. Preview of each messaging section
  2. finance_node — VA-07 congressional, $75,000 budget
  2a. Structured data schema check
  2b. FEC data fetch confirmation
  2c. Print full budget narrative
  3. export_node — text format
  4. export_node — markdown format
  5. export_node — docx format (file-existence check)
  6. export_node — xlsx format (file-existence check)

Run from project root:
  python -m chat.tests.test_agents
"""

from dotenv import load_dotenv
load_dotenv()  # must be before any import that reads env vars

import logging
import os
import sys

# -- Path setup ---------------------------------------------------------------
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

# -- Logging ------------------------------------------------------------------
logging.basicConfig(
    level=logging.WARNING,
    format="  [LOG] %(message)s",
    stream=sys.stdout,
)
for noisy in ("httpx", "httpcore", "openai", "urllib3", "requests", "langchain"):
    logging.getLogger(noisy).setLevel(logging.WARNING)

# -- Test helpers -------------------------------------------------------------

_GREEN  = "\033[32m"
_RED    = "\033[31m"
_YELLOW = "\033[33m"
_RESET  = "\033[0m"

_results: list = []


def check(name: str, actual, expected, *, note: str = "") -> bool:
    passed = actual == expected
    tag = f"{_GREEN}PASS{_RESET}" if passed else f"{_RED}FAIL{_RESET}"
    _results.append((name, passed))
    print(f"  [{tag}] {name}")
    if not passed:
        print(f"         expected : {expected!r}")
        print(f"         actual   : {actual!r}")
    elif note:
        print(f"         {_YELLOW}{note}{_RESET}")
    return passed


def check_true(name: str, value: bool, *, note: str = "") -> bool:
    return check(name, bool(value), True, note=note)


def section(title: str):
    print(f"\n{'-' * 64}")
    print(f"  {title}")
    print("-" * 64)


def skip(name: str, reason: str):
    print(f"  [{_YELLOW}SKIP{_RESET}] {name}")
    print(f"         {reason}")
    _results.append((name, None))


def info(label: str, value=""):
    if value:
        print(f"  [INFO] {label}: {value}")
    else:
        print(f"  [INFO] {label}")


def warn(msg: str):
    print(f"  [{_YELLOW}WARN{_RESET}] {msg}")


def preview(label: str, text: str, chars: int = 500):
    """Print a truncated preview of a long string."""
    trimmed = text.strip()[:chars]
    ellipsis = "..." if len(text.strip()) > chars else ""
    print(f"\n  -- {label} --")
    for line in (trimmed + ellipsis).splitlines():
        print(f"  {line}")


# =============================================================================
# Mock AgentState data — VA-07, young voter outreach
# =============================================================================

# Real VA-07 top precincts from the precinct pipeline (Henrico + Chesterfield)
MOCK_PRECINCTS = [
    {
        "precinct_geoid":     "51087-315-Tucker",
        "precinct_name":      "Tucker",
        "vap":                6917,
        "total_cvap":         5840,
        "black":              1620,
        "hispanic":           530,
        "white":              3340,
        "approximate_boundary": False,
    },
    {
        "precinct_geoid":     "51041-503-Midlothian",
        "precinct_name":      "Midlothian",
        "vap":                6744,
        "total_cvap":         5920,
        "black":              980,
        "hispanic":           410,
        "white":              4220,
        "approximate_boundary": False,
    },
    {
        "precinct_geoid":     "51041-414-Crenshaw",
        "precinct_name":      "Crenshaw",
        "vap":                6203,
        "total_cvap":         5540,
        "black":              1890,
        "hispanic":           620,
        "white":              2730,
        "approximate_boundary": False,
    },
    {
        "precinct_geoid":     "51087-105-Hermitage",
        "precinct_name":      "Hermitage",
        "vap":                5989,
        "total_cvap":         5120,
        "black":              2410,
        "hispanic":           580,
        "white":              1880,
        "approximate_boundary": False,
    },
    {
        "precinct_geoid":     "51041-207-Chippenham",
        "precinct_name":      "Chippenham",
        "vap":                5945,
        "total_cvap":         5200,
        "black":              1540,
        "hispanic":           490,
        "white":              2960,
        "approximate_boundary": False,
    },
]

MOCK_PRECINCTS_ENTRY = {
    "agent":             "precincts",
    "state_fips":        "51",
    "district_type":     "congressional",
    "district_id":       "5107",
    "precincts":         MOCK_PRECINCTS,
    "precinct_count":    295,
    "data_quality_note": None,
}

# Realistic win-number data for VA-07 2026 midterm projection
MOCK_WIN_NUMBER_ENTRY = {
    "agent":                "win_number",
    "state_fips":           "51",
    "district_type":        "congressional",
    "district_id":          "5107",
    "target_year":          2026,
    "win_number":           154214,
    "projected_turnout":    296565,
    "voter_universe_cvap":  522000,
    "avg_turnout_pct":      0.568,
    "victory_margin":       0.52,
    "historical_context":   "Based on 2018 and 2022 midterm cycles",
    "data_notes": (
        "Midterm projection based on [2018, 2022]. "
        "2022 cycle data not yet integrated — projection may skew high."
    ),
}

# Mock research findings — formatted to match researcher.py memo headers
# Uses real publication dates so the recency sort and recency notes work
MOCK_RESEARCH_RESULTS = [
    (
        "--- MEMO FROM SOURCE: Young-Men-Focus-Groups-Report_final.pdf"
        " | DATE: 2025-06-01 ---\n"
        "Focus groups with young men ages 18-35 in Virginia suburban districts reveal "
        "three dominant concerns: economic stability and job quality, housing affordability, "
        "and a perceived lack of candidates who speak to them directly. Participants across "
        "racial groups consistently said they would turn out if they felt a candidate had "
        "concrete plans rather than talking points. Canvassers who opened with economic "
        "questions saw 40% longer conversations than those who led with party affiliation. "
        "Text outreach with personalized first names drove a 22% callback rate compared to "
        "8% for generic blasts. Door-knocking in the evening (5-8pm) was significantly more "
        "effective in this demographic than weekend morning canvass shifts.\n"
    ),
    (
        "--- MEMO FROM SOURCE: HIT_Headcount_Public-Memo.pdf"
        " | DATE: 2024-12-17 ---\n"
        "A headcount analysis of Virginia's 7th Congressional District identifies 295 "
        "polling precincts with a combined voting-age population of approximately 875,000. "
        "Approximately 57% of eligible voters participated in the 2022 midterm. Black voters "
        "represent 28% of the CVAP universe in the Henrico and Chesterfield County precincts "
        "that anchor the district. Hispanic CVAP is growing — up 11% since 2020 — and remains "
        "under-contacted by both major parties. Young voters (18-29) make up 18% of the VAP "
        "but only 11% of 2022 midterm participants, representing the largest mobilization gap. "
        "Turnout models suggest a 4-point improvement in this cohort would shift the district "
        "by approximately 6,000 net votes.\n"
    ),
]

QUERY_MESSAGING  = (
    "Build a young voter outreach messaging program for Virginia's 7th Congressional District."
)
QUERY_FINANCE    = (
    "What will it cost to run a young voter field program in Virginia's 7th Congressional "
    "District? We have a $75,000 budget available."
)
QUERY_EXPORT     = (
    "Build a complete program plan for young voter outreach in Virginia's 7th Congressional District."
)

# =============================================================================
# 0. Pre-flight
# =============================================================================

section("0 - Pre-flight checks")

has_openai = bool(os.environ.get("OPENAI_API_KEY"))
has_fec    = bool(os.environ.get("FEC_API_KEY"))

if has_openai:
    check_true("OPENAI_API_KEY is set", True)
else:
    skip("OPENAI_API_KEY is set", "Not set — all agent tests will be skipped")

if has_fec:
    check_true("FEC_API_KEY is set (finance historical mode)", True)
else:
    info("FEC_API_KEY not set — finance_node will run in unit-cost-only mode")

try:
    from docx import Document as _DocxDoc
    has_docx = True
    check_true("python-docx is installed (required for docx export)", True)
except ImportError:
    has_docx = False
    skip("python-docx installed", "pip install python-docx")

try:
    import openpyxl as _openpyxl
    has_openpyxl = True
    check_true("openpyxl is installed (required for xlsx export)", True)
except ImportError:
    has_openpyxl = False
    skip("openpyxl installed", "pip install openpyxl")


# =============================================================================
# 1. messaging_node
# =============================================================================

section("1 - messaging_node — VA-07 young voter outreach")

messaging_research_out: list = []

if not has_openai:
    skip("messaging_node", "OPENAI_API_KEY not set")
else:
    from chat.agents.messaging import messaging_node, FORMAT_LABELS

    print("  Constructing AgentState with mock precinct demographics + research findings...")
    print("  Making LLM call (gpt-4o) to generate all five messaging formats...")
    print()

    messaging_state = {
        "query":           QUERY_MESSAGING,
        "org_namespace":   "general",
        "structured_data": [MOCK_PRECINCTS_ENTRY],
        "research_results": MOCK_RESEARCH_RESULTS,
        "active_agents":   ["researcher"],
        "errors":          [],
    }

    try:
        msg_result = messaging_node(messaging_state)
        msg_ok = True
    except Exception as e:
        msg_result = {}
        msg_ok = False
        warn(f"messaging_node raised: {e}")

    check_true("messaging_node returned without exception", msg_ok)

    if msg_ok:
        msg_errors = msg_result.get("errors", [])
        check_true(
            "messaging_node returned no errors",
            len(msg_errors) == 0,
            note=str(msg_errors) if msg_errors else "",
        )

        messaging_research_out = msg_result.get("research_results", [])
        check_true(
            "messaging_node returned at least one output",
            len(messaging_research_out) > 0,
            note=f"{len(messaging_research_out)} messaging sections returned",
        )

        # -- 1a: Section label checks -----------------------------------------
        section("1a - Messaging section labels present in output")

        found_labels = []
        for label in FORMAT_LABELS.values():
            present = any(label in m for m in messaging_research_out)
            check_true(f'Section "{label}" present in output', present)
            if present:
                found_labels.append(label)

        info(f"Sections found", f"{len(found_labels)}/5")

        # -- 1b: Section previews ---------------------------------------------
        section("1b - Messaging output previews (first 400 chars per section)")

        for i, memo in enumerate(messaging_research_out, 1):
            # Extract label from the memo header line
            header_end = memo.find("\n")
            header = memo[:header_end].strip() if header_end > 0 else f"Section {i}"
            body   = memo[header_end:].strip() if header_end > 0 else memo
            preview(header, body, chars=400)


# =============================================================================
# 2. finance_node — VA-07, $75k budget
# =============================================================================

section("2 - finance_node — VA-07 congressional, $75,000 budget")

finance_structured_out: list = []
finance_research_out:   list = []

if not has_openai:
    skip("finance_node", "OPENAI_API_KEY not set")
else:
    from chat.agents.finance_agent import finance_node

    print("  Geographic context provided via structured_data (no LLM extraction needed).")
    print("  Budget extracted from query via LLM. FEC API call for 2018+2022 cycles...")
    print()

    finance_state = {
        "query":           QUERY_FINANCE,
        "org_namespace":   "general",
        # precincts entry carries state_fips/district_type/district_id →
        # finance_node uses it directly, skipping LLM geo extraction
        "structured_data": [MOCK_PRECINCTS_ENTRY],
        "research_results": [],
        "active_agents":   ["precincts"],
        "errors":          [],
    }

    try:
        fin_result = finance_node(finance_state)
        fin_ok = True
    except Exception as e:
        fin_result = {}
        fin_ok = False
        warn(f"finance_node raised: {e}")

    check_true("finance_node returned without exception", fin_ok)

    if fin_ok:
        # Non-fatal FEC errors are expected when FEC_API_KEY is absent or data is sparse
        fin_errors = fin_result.get("errors", [])
        if fin_errors:
            for e in fin_errors:
                warn(f"finance_node non-fatal error: {e}")

        finance_structured_out = fin_result.get("structured_data", [])
        finance_research_out   = fin_result.get("research_results", [])

        # -- 2a: Structured data schema ---------------------------------------
        section("2a - Finance structured_data schema")

        check_true(
            "structured_data contains a finance entry",
            len(finance_structured_out) > 0,
            note=f"{len(finance_structured_out)} entry/entries",
        )

        if finance_structured_out:
            fin_entry = finance_structured_out[0]

            check_true(
                'Finance entry has "agent" == "finance"',
                fin_entry.get("agent") == "finance",
                note=str(fin_entry.get("agent")),
            )
            check_true(
                'Finance entry has "unit_costs" dict',
                isinstance(fin_entry.get("unit_costs"), dict),
            )
            check_true(
                'Finance entry has "mode" field',
                fin_entry.get("mode") in ("historical", "hybrid", "unit_cost_only"),
                note=f"mode = {fin_entry.get('mode')}",
            )

            budget_prog = fin_entry.get("budget_program")
            check_true(
                "budget_program populated (budget was extracted from query)",
                isinstance(budget_prog, dict) and len(budget_prog) > 0,
                note=f"tactics: {list(budget_prog.keys()) if budget_prog else 'none'}",
            )

            # -- 2b: FEC data check -------------------------------------------
            section("2b - FEC historical data check")

            mode      = fin_entry.get("mode", "unit_cost_only")
            sampled   = fin_entry.get("fec_candidates_sampled", 0)
            full_est  = fin_entry.get("full_program_estimate")
            info("Finance mode", mode)
            info("FEC candidates sampled", sampled)

            if mode in ("historical", "hybrid"):
                check_true(
                    "FEC candidates sampled > 0 (historical data found)",
                    sampled > 0,
                    note=f"{sampled} candidates across comparable midterm cycles",
                )
                check_true(
                    "full_program_estimate populated from FEC data",
                    isinstance(full_est, dict) and full_est.get("total", 0) > 0,
                    note=f"Total: ${full_est.get('total', 0):,.0f}" if full_est else "none",
                )
            else:
                _results.append(("FEC candidates sampled > 0", None))
                print(
                    f"  [{_YELLOW}INFO{_RESET}] mode=unit_cost_only — FEC data unavailable "
                    "(expected when FEC_API_KEY is absent or district data is sparse)."
                )

        # -- 2c: Print full budget narrative ----------------------------------
        section("2c - Finance budget narrative (full output)")

        if finance_research_out:
            for memo in finance_research_out:
                for line in memo.splitlines():
                    print(f"  {line}")
        else:
            warn("No narrative output in research_results.")


# =============================================================================
# Assemble full export state from all agent outputs
# =============================================================================

# Merge structured_data and research_results from all agents
assembled_structured = (
    [MOCK_PRECINCTS_ENTRY, MOCK_WIN_NUMBER_ENTRY]
    + finance_structured_out
)
assembled_research = (
    MOCK_RESEARCH_RESULTS
    + messaging_research_out
    + finance_research_out
)
# All six plan agents active → export_node triggers full plan synthesis (is_plan=True)
PLAN_AGENTS_ACTIVE = [
    "researcher", "election_results", "win_number",
    "precincts", "messaging", "finance",
]


def _make_export_state(output_format: str) -> dict:
    return {
        "query":           QUERY_EXPORT,
        "org_namespace":   "general",
        "output_format":   output_format,
        "structured_data": assembled_structured,
        "research_results": assembled_research,
        "active_agents":   PLAN_AGENTS_ACTIVE,
        "errors":          [],
    }


# =============================================================================
# 3. export_node — text
# =============================================================================

section("3 - export_node — output_format='text'")

if not has_openai:
    skip("export_node (text)", "OPENAI_API_KEY not set")
else:
    from chat.agents.export import export_node

    print("  Synthesizing full political program plan (gpt-4o)...")
    print("  active_agents includes all 6 plan agents -> is_plan=True")
    print()

    try:
        text_result = export_node(_make_export_state("text"))
        text_ok = True
    except Exception as e:
        text_result = {}
        text_ok = False
        warn(f"export_node (text) raised: {e}")

    check_true("export_node (text) returned without exception", text_ok)

    if text_ok:
        final_ans = text_result.get("final_answer", "")
        check_true(
            "final_answer is a non-empty string",
            isinstance(final_ans, str) and len(final_ans) > 100,
            note=f"{len(final_ans):,} chars",
        )
        check_true(
            "generated_file_path is NOT set for text format",
            text_result.get("generated_file_path") is None,
        )
        export_errors = text_result.get("errors", [])
        if export_errors:
            for e in export_errors:
                warn(f"export error: {e}")

        preview("Synthesized output (text) — first 800 chars", final_ans, chars=800)


# =============================================================================
# 4. export_node — markdown
# =============================================================================

section("4 - export_node — output_format='markdown'")

if not has_openai:
    skip("export_node (markdown)", "OPENAI_API_KEY not set")
else:
    print("  Synthesizing (gpt-4o) — markdown and text use the same handler...")
    print()

    try:
        md_result = export_node(_make_export_state("markdown"))
        md_ok = True
    except Exception as e:
        md_result = {}
        md_ok = False
        warn(f"export_node (markdown) raised: {e}")

    check_true("export_node (markdown) returned without exception", md_ok)

    if md_ok:
        final_ans_md = md_result.get("final_answer", "")
        check_true(
            "final_answer is a non-empty string",
            isinstance(final_ans_md, str) and len(final_ans_md) > 100,
            note=f"{len(final_ans_md):,} chars",
        )
        check_true(
            "generated_file_path is NOT set for markdown format",
            md_result.get("generated_file_path") is None,
        )
        check_true(
            "Markdown output contains at least one H2 section header",
            "## " in final_ans_md,
        )
        preview("Synthesized output (markdown) — first 800 chars", final_ans_md, chars=800)


# =============================================================================
# 5. export_node — docx
# =============================================================================

section("5 - export_node — output_format='docx'")

if not has_openai:
    skip("export_node (docx)", "OPENAI_API_KEY not set")
elif not has_docx:
    skip("export_node (docx)", "python-docx not installed")
else:
    print("  Synthesizing + building Word document (python-docx)...")
    print("  Precinct table and win-number table injected from structured_data...")
    print()

    try:
        docx_result = export_node(_make_export_state("docx"))
        docx_ok = True
    except Exception as e:
        docx_result = {}
        docx_ok = False
        warn(f"export_node (docx) raised: {e}")

    check_true("export_node (docx) returned without exception", docx_ok)

    if docx_ok:
        docx_errors = docx_result.get("errors", [])
        if docx_errors:
            for e in docx_errors:
                warn(f"export error: {e}")

        docx_path = docx_result.get("generated_file_path")
        check_true(
            "generated_file_path is set for docx format",
            docx_path is not None,
            note=str(docx_path),
        )
        if docx_path:
            exists = os.path.isfile(docx_path)
            check_true("docx file exists on disk", exists, note=docx_path)
            if exists:
                size_kb = os.path.getsize(docx_path) / 1024
                check_true(
                    "docx file size > 0 bytes",
                    size_kb > 0,
                    note=f"{size_kb:.1f} KB",
                )
                info("docx file path", docx_path)
                info("docx file size", f"{size_kb:.1f} KB")

        check_true(
            "final_answer populated for docx format",
            isinstance(docx_result.get("final_answer", ""), str)
            and len(docx_result.get("final_answer", "")) > 50,
        )


# =============================================================================
# 6. export_node — xlsx
# =============================================================================

section("6 - export_node — output_format='xlsx'")

if not has_openai:
    skip("export_node (xlsx)", "OPENAI_API_KEY not set")
elif not has_openpyxl:
    skip("export_node (xlsx)", "openpyxl not installed")
else:
    print("  Synthesizing + building Excel workbook (openpyxl)...")
    print("  Three sheets: Precinct Targets, Win Number, Budget Estimate...")
    print()

    try:
        xlsx_result = export_node(_make_export_state("xlsx"))
        xlsx_ok = True
    except Exception as e:
        xlsx_result = {}
        xlsx_ok = False
        warn(f"export_node (xlsx) raised: {e}")

    check_true("export_node (xlsx) returned without exception", xlsx_ok)

    if xlsx_ok:
        xlsx_errors = xlsx_result.get("errors", [])
        if xlsx_errors:
            for e in xlsx_errors:
                warn(f"export error: {e}")

        xlsx_path = xlsx_result.get("generated_file_path")
        check_true(
            "generated_file_path is set for xlsx format",
            xlsx_path is not None,
            note=str(xlsx_path),
        )
        if xlsx_path:
            exists = os.path.isfile(xlsx_path)
            check_true("xlsx file exists on disk", exists, note=xlsx_path)
            if exists:
                size_kb = os.path.getsize(xlsx_path) / 1024
                check_true(
                    "xlsx file size > 0 bytes",
                    size_kb > 0,
                    note=f"{size_kb:.1f} KB",
                )
                info("xlsx file path", xlsx_path)
                info("xlsx file size", f"{size_kb:.1f} KB")

                # Verify sheet names using openpyxl directly
                import openpyxl
                wb = openpyxl.load_workbook(xlsx_path)
                expected_sheets = {"Precinct Targets", "Win Number", "Budget Estimate"}
                actual_sheets   = set(wb.sheetnames)
                check_true(
                    "xlsx has all three expected sheets",
                    expected_sheets.issubset(actual_sheets),
                    note=f"Sheets: {sorted(actual_sheets)}",
                )

                # Verify precinct data landed in Sheet 1
                ws1 = wb["Precinct Targets"]
                precinct_rows = ws1.max_row
                check_true(
                    "Precinct Targets sheet has data rows (header + precincts)",
                    precinct_rows > 1,
                    note=f"{precinct_rows} rows (including header)",
                )

        check_true(
            "final_answer populated for xlsx format",
            isinstance(xlsx_result.get("final_answer", ""), str)
            and len(xlsx_result.get("final_answer", "")) > 10,
        )


# =============================================================================
# Summary
# =============================================================================

section("Summary")

passed  = sum(1 for _, ok in _results if ok is True)
failed  = sum(1 for _, ok in _results if ok is False)
skipped = sum(1 for _, ok in _results if ok is None)
total   = len(_results)

print(f"\n  Total : {total}")
print(f"  {_GREEN}Passed{_RESET}: {passed}")
if failed:
    print(f"  {_RED}Failed{_RESET}: {failed}")
if skipped:
    print(f"  {_YELLOW}Skipped{_RESET}: {skipped}")
print()

if failed:
    print(f"  {_RED}FAILED TESTS:{_RESET}")
    for name, ok in _results:
        if ok is False:
            print(f"    x {name}")
    sys.exit(1)
else:
    print(
        f"  {_GREEN}All checks passed.{_RESET}" if not skipped
        else f"  {_GREEN}All non-skipped checks passed.{_RESET}"
    )
    sys.exit(0)
